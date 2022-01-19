from openpyxl import Workbook

arquivo_excel = Workbook()
planilha1 = arquivo_excel.active
planilha1.title = 'Produtos'
planilha2 = arquivo_excel.create_sheet("Clientes")
planilha3 = arquivo_excel.create_sheet("Mensalistas")
print(arquivo_excel.sheetnames)
planilha1['A1'] = 'Nome Produto'
planilha1['B1'] = 'Preço Produto'
planilha1['C1'] = 'Quantidade'

produtos = [("Arroz", 17.95, 250),
            ("Feijão", 8.90, 500)
            ]
for linha in produtos:
    planilha1.append(linha)
planilha1['D2'] = '=C2 * B2'
planilha1['D3'] = '=C3 * B3'
planilha2.cell(row=1, column=1, value='Rafael')

#Contagem de produtos = def/while/if
max_linha = planilha1.max_row
max_coluna = planilha1.max_column
for i in range(1, max_linha + 1):
    for j in range(2, max_coluna + 1):
        print(planilha1.cell(row=i, column=j).value, end='-')
arquivo_excel.save("Relátorio.xlsx")