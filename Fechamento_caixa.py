from openpyxl import load_workbook


class Fechamento_Caixa:

    def fechamentoCaixa(self):
        arquivo_excel = load_workbook('Caixa_Screen.xlsx')
        caixa_screen = arquivo_excel.active
        caixa_screen = arquivo_excel.create_sheet("Produtos")
        recebido = float(input('Qual o o valor total recebido? R$ '))
        saida = float(input('Qual o valor total gasto? R$ '))
        faturamento = recebido - saida

        caixa_screen['B4'] = recebido
        caixa_screen['C4'] = saida
        caixa_screen['D4'] = faturamento

        arquivo_excel.save('Caixa-Screen.xlsx')


if __name__ == '__main__':
    fechamento_do_caixa = Fechamento_Caixa()
